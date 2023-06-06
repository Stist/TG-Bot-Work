import telebot, random, time, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from telebot.types import User

bot = telebot.TeleBot("6122034197:AAEOKGrsxUGeWqCHbun3Vfgd4hKrhuzwbwE")

file_path = r"/home/administrator/Рабочий стол/TGBot_Final/JC_Table.xlsx"
wb = load_workbook(filename=file_path)
ws = wb['Menu']
worksheet = wb.active
wb.close() 

user_ID = ""   #   Переменная для хранения ID пользователя

message_del = ""   #   Переменная для хранения значения последнего сообщения

Flag_problems = False   #   Флаг, работает во время заведения заявки

Flag_my_problems = False   #   Флаг, работает во время просмотра своих заявок

Flag_debug = False   #   Флаг, для проверки на спам

Flag_add_SV = False

Flag_add_admin = False

Flag_add = False

Flag_del_SV = False

Flag_del_admin = False

cell_copy = ""   #   Переменная для хранения cell в тех случаях, когда необходимо*

copy_call_or_message = ""

last_message = ""   #   Переменная для правильной отработки условий*

last_location = ""   #   Переменная для хранения местоположения при удалении заявки

last_button_cells = ""

location_last_empty = ""

text_end_report = ""   #   Переменная описания новой заявки

message_list = []   #   Список ID сообщений, который отображается на экране у пользователя

message_global = ""   #   Копия message для всех функций

message_warning = ""

user_link = ""  #   Ссылка пользователя на телеграм

for_user_link = ""  #   Ссылка пользователя на телеграм

undel_mes = ""   #   Костыль, поскольку удаление этот сообщения через список сообщений, вызывает ошибку

last_message_times = {} # Словарь для хранения времени последнего сообщения от каждого пользователя


#   Переменные для составлени заявки
Kind = ""   
Team = ""
Name = ""
Sip = ""
Of_Ud = ""
How_managed = ""
decribtion = ""
my_proj_men = ""


#   Переменные для хранения ФИ и ID ответственного
my_IT = worksheet.cell(51,3).value
id_my_IT = worksheet.cell(52,3).value


#   Переменные для добавление УЗ СВ или Адммина
Name_SV = ""
id_SV = ""
link_SV = ""
proj_SV = ""



###     Создание/Обновление списков     ###



#   Создаём список забаненых пользователей
ban_list = []
with open("ban.txt", 'r') as file:
    for line in file:
        user_id = line.strip()  # Удаляем символы перевода строки и пробелы
        ban_list.append(user_id)  # Добавляем ID в список
#   Обновляем список забаненых пользователей
def ban_list_refresh():
    global ban_list
    ban_list = []
    with open("ban.txt", 'r') as file:
        for line in file:
            user_id = line.strip()  # Удаляем символы перевода строки и пробелы
            ban_list.append(user_id)  # Добавляем ID в список


#   Создаём список пользователей-супервайзеров
acc_list = []
for row in worksheet.iter_rows(min_row=45, max_row=45):
    for cell in row:
        if cell.column > 2:
            if(cell.value is None):
                break
            else:
                acc_list.append(cell.value)
#   Обновляем список пользователей-супервайзеров
def acc_list_refresh():
    global acc_list
    acc_list = []
    for row in worksheet.iter_rows(min_row=45, max_row=45):
        for cell in row:
            if cell.column > 2:
                if(cell.value is None):
                    break
                else:
                    acc_list.append(cell.value)


#   Создаём список пользователей-администраторов
admin_list = []
for row in worksheet.iter_rows(min_row=52, max_row=52):
    for cell in row:
        if cell.column > 6:
            if(cell.value is None):
                break
            else:
                admin_list.append(cell.value)
#   Обновляем список пользователей-администраторов
def admin_list_refresh():
    global admin_list
    admin_list = []
    for row in worksheet.iter_rows(min_row=52, max_row=52):
        for cell in row:
            if cell.column > 6:
                if(cell.value is None):
                    break
                else:
                    admin_list.append(cell.value)



###     Добавление и удаление СВ / Администраторов     ###



#   Меню Администраторов
def Menu_AdminMenu():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV    
    button_cells = []


    button_cells.append(telebot.types.InlineKeyboardButton("Добавить Супервайзера", callback_data="[add_SV]"))
    button_cells.append(telebot.types.InlineKeyboardButton("Удалить Супервайзера", callback_data="[del_SV]"))
    button_cells.append(telebot.types.InlineKeyboardButton("Добавить Администратора", callback_data="[add_admin]"))
    button_cells.append(telebot.types.InlineKeyboardButton("Удалить Администратора", callback_data="[del_admin]"))
    button_cells.append(telebot.types.InlineKeyboardButton("-", callback_data="[none]"))
    button_cells.append(telebot.types.InlineKeyboardButton("Вернуться", callback_data="[back_menu]"))

    last_button_cells = button_cells    # Переменная для вывода последнего меню
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*button_cells)
    return inline_keyboard

#   Заход - Выход (В таблицу)
#   Добавление нового СВ
def Menu_AdminMenu_AddSV():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV
    
    wb = load_workbook(filename=file_path)
    worksheet = wb.active

    for row in worksheet.iter_rows(min_row=44, max_row=44):
        for cell in row:
            if cell.column > 2:
                if(cell.value == None):
                    ws.cell(row=44, column=cell.column, value=Name_SV)

                    ws.cell(row=45, column=cell.column, value=id_SV)

                    ws.cell(row=46, column=cell.column, value=link_SV)

                    ws.cell(row=47, column=cell.column, value=proj_SV)
                    
                    break   #   Нашли - Выполнили - Выходим

    Flag_add_SV = False
    wb.save(file_path)
    wb.close()

    bot.send_message(id_my_IT, f"✅ Добавлен новый сотрудник в группу СВ!\n\nСотрудник: \"{Name_SV}\"\n\nКем: {user_ID}\n")
    text = bot.send_message(user_ID, f"✅ Сотрудник \"{Name_SV}\" добавлен в группу СВ!")
    
    file = open('log.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
    else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
    file.close()

    Clear_Nums_And_Go_To_Menu()

#   Заход - Выход (В таблицу)
#   Добавление нового Администратора (В группу СВ тоже)
def Menu_AdminMenu_AddAdmin():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    wb = load_workbook(filename=file_path)
    worksheet = wb.active

    for row in worksheet.iter_rows(min_row=51, max_row=51):
        for cell in row:
            if cell.column > 6:
                if(cell.value == None):
                    ws.cell(row=51, column=cell.column, value=Name_SV)   #   Удаляю ID User

                    ws.cell(row=52, column=cell.column, value=id_SV)
                    break

                elif(cell.value == Name_SV):
                    break   #   Нашли - Выполнили - Выходим

    for row in worksheet.iter_rows(min_row=45, max_row=45):
        for cell in row:
            if cell.column > 2:
                if(cell.value == None):
                    ws.cell(row=44, column=cell.column, value=Name_SV)   #   Удаляю ID User

                    ws.cell(row=45, column=cell.column, value=id_SV)   #   Удаляю описание проблемы

                    ws.cell(row=46, column=cell.column, value=link_SV)   #   Удаляю ID IT

                    ws.cell(row=47, column=cell.column, value=proj_SV)   #   Удаляю ID Менеджера
                    break

                elif(cell.value == id_SV):
                    break   #   Нашли - Выполнили - Выходим

    Flag_add_admin = False
    wb.save(file_path)
    wb.close()
    bot.send_message(id_my_IT, f"✅ Добавлен новый сотрудник в группу Администраторы!\n\nСотрудник: \"{Name_SV}\"\n\nКем: {user_ID}\n")
    text = bot.send_message(user_ID, f"✅ Сотрудник \"{Name_SV}\" добавлен в группу Администраторы!")

    file = open('log.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
    else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
    file.close()

    Clear_Nums_And_Go_To_Menu()

#   Заход - Выход (В таблицу)
#   Удаление СВ  (Из группы Администраторов тоже)
def Menu_AdminMenu_DelSV(del_id):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_del_SV
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    wb = load_workbook(filename=file_path)
    worksheet = wb.active

    Flag_del_SV = False

    for row in worksheet.iter_rows(min_row=45, max_row=45):
        for cell in row:
            if cell.column > 2:
                if(cell.value == del_id):

                    ws.cell(row=44, column=cell.column, value="")   #   Удаляю ID User

                    ws.cell(row=45, column=cell.column, value="")   #   Удаляю описание проблемы

                    ws.cell(row=46, column=cell.column, value="")   #   Удаляю ID IT

                    ws.cell(row=47, column=cell.column, value="")   #   Удаляю ID Менеджера
                    

                    cell = worksheet.cell(cell.row, cell.column+1)

                    while cell.value is not None:
                        ws.cell(row=44, column=cell.column-1, value=worksheet.cell(44, cell.column).value)   #   Удаляю ID User
                        ws.cell(row=45, column=cell.column-1, value=worksheet.cell(45, cell.column).value)   #   Удаляю описание проблемы
                        ws.cell(row=46, column=cell.column-1, value=worksheet.cell(46, cell.column).value)   #   Удаляю ID IT
                        ws.cell(row=47, column=cell.column-1, value=worksheet.cell(47, cell.column).value)   #   Удаляю ID Менеджера

                        
                        ws.cell(row=44, column=cell.column, value="")   #   Удаляю ID User
                        ws.cell(row=45, column=cell.column, value="")   #   Удаляю описание проблемы
                        ws.cell(row=46, column=cell.column, value="")   #   Удаляю ID IT
                        ws.cell(row=47, column=cell.column, value="")   #   Удаляю ID Менеджера

                        if cell.column == 200:
                            break
                        cell = worksheet.cell(cell.row, cell.column+1)
                    
                    Menu_AdminMenu_DelAdmin(del_id)

                    wb.save(file_path)
                    wb.close()

                    Del_All_Message()   #   Удаляю все сообщения
                    text = bot.send_message(user_ID, f"🚫 Супервайзер с ID {del_id} был удалён!")

                    file = open('log.txt', 'a')
                    now_time = datetime.datetime.now()
                    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
                    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
                    else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
                    file.close()

                    Clear_Nums_And_Go_To_Menu()
                    break

                elif(cell.value == None):

                    wb.save(file_path)
                    wb.close()

                    Del_All_Message()   #   Удаляю все сообщения
                    bot.send_message(user_ID, f"⚠️ Супервайзер с ID {del_id} не найден!")
                    Clear_Nums_And_Go_To_Menu()
                    break

#   Заход - Выход (В таблицу)
#   Удаление Администратора
def Menu_AdminMenu_DelAdmin(del_id):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_del_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    wb = load_workbook(filename=file_path)
    worksheet = wb.active
    
    Flag_del_admin = False
    
    for row in worksheet.iter_rows(min_row=52, max_row=52):
        for cell in row:
            if cell.column > 6:
                if(cell.value == del_id):

                    ws.cell(row=51, column=cell.column, value="")   #   Удаляю ID User

                    ws.cell(row=52, column=cell.column, value="")   #   Удаляю описание проблемы
                    

                    cell = worksheet.cell(cell.row, cell.column+1)

                    while cell.value is not None:
                        ws.cell(row=51, column=cell.column-1, value=worksheet.cell(51, cell.column).value)   #   Удаляю ID User
                        ws.cell(row=52, column=cell.column-1, value=worksheet.cell(52, cell.column).value)   #   Удаляю описание проблемы

                        
                        ws.cell(row=51, column=cell.column, value="")   #   Удаляю ID User
                        ws.cell(row=52, column=cell.column, value="")   #   Удаляю описание проблемы

                        if cell.column == 100:
                            break
                        cell = worksheet.cell(cell.row, cell.column+1)

                    wb.save(file_path)
                    wb.close()

                    Del_All_Message()   #   Удаляю все сообщения
                    text = bot.send_message(user_ID, f"🚫 Администратор с ID {del_id} был удалён!")
                    
                    file = open('log.txt', 'a')
                    now_time = datetime.datetime.now()
                    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
                    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
                    else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
                    file.close()


                    message_list.append(message_del.message_id)
                    Clear_Nums_And_Go_To_Menu()
                    break

                elif(cell.value == None):

                    wb.save(file_path)
                    wb.close()

                    Del_All_Message()   #   Удаляю все сообщения
                    bot.send_message(user_ID, f"⚠️ Администратор с ID {del_id} не найден!")
                    message_list.append(message_del.message_id)
                    Clear_Nums_And_Go_To_Menu()
                    break



###     Меню - Прочее     ###



#   Функция (Побочная) поиска/вывода ID пользователя
def Menu_Enother_TakeID():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    Del_All_Message()
    message_del = bot.send_message(user_ID, f"📎 Ваш ID: {user_ID}")
    message_list.append(message_del.message_id)
    message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_Movement("[enother]"))
    message_list.append(message_del.message_id)

#   Функция вывода Link пользователя
def Menu_Enother_TakeLink():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, for_user_link
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV
    
    Del_All_Message()
    message_del = bot.send_message(user_ID, f"📎 Ваша ссылка на telegram: https://t.me/{for_user_link}")
    message_list.append(message_del.message_id)
    message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_Movement("[enother]"))
    message_list.append(message_del.message_id)



###     Меню - Заявки     ###



#   Функция создания меню для модуля "Заявки"
def Menu_ReportsMenu():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV
    button_cells = []


    for y in acc_list:
        if(user_ID == y):
            button_cells.append(telebot.types.InlineKeyboardButton("Создать заявку", callback_data="[create_report]"))
            button_cells.append(telebot.types.InlineKeyboardButton("Мои заявки", callback_data="[my_report]"))


    for y in admin_list:
        if(user_ID == y):
            button_cells.append(telebot.types.InlineKeyboardButton("Все заявки", callback_data="[all_reports]"))

    button_cells.append(telebot.types.InlineKeyboardButton("-", callback_data="[none]"))
    button_cells.append(telebot.types.InlineKeyboardButton("Вернуться", callback_data="[back_menu]"))

    last_button_cells = button_cells    # Переменная для вывода последнего меню
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*button_cells)
    return inline_keyboard

#   Заход - Выход (В таблицу)
#   Функция вывода списка своих заявок
def Menu_ReportsMenu_MyReports():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV


    wb = load_workbook(filename=file_path)
    ws = wb['Menu']
    worksheet = wb.active
    my_report_list = []
    Flag_my_problems = True

    for row in worksheet.iter_rows(min_row=60, max_row=60):
        for cell in row:
            if cell.column > 2:
                if((cell.value is not None) and (user_ID == worksheet.cell(cell.row-4, cell.column).value)):    #   Проверка на то, что заявка именно его.
                    my_report_list.append(telebot.types.InlineKeyboardButton(cell.value, callback_data=cell.value))
                    
                elif cell.value is None:
                    break
    
    wb.close()

    if(my_report_list == []):
        my_report_list.append(telebot.types.InlineKeyboardButton("(Пусто)", callback_data="[none]"))
    my_report_list.append(telebot.types.InlineKeyboardButton("-", callback_data="[none]"))
    my_report_list.append(telebot.types.InlineKeyboardButton("Вернуться", callback_data="[back_menu_my_reports]"))
    
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*my_report_list)
    return inline_keyboard

#   Заход - Выход (В таблицу)
#   Функция вывода списка ВСЕХ заявок
def Menu_ReportsMenu_AllReports():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
        
    wb = load_workbook(filename=file_path)
    ws = wb['Menu']
    worksheet = wb.active
    my_report_list = []
    Flag_my_problems = True

    for row in worksheet.iter_rows(min_row=60, max_row=60):
        for cell in row:
            if cell.column > 2:
                if(cell.value is not None):
                    my_report_list.append(telebot.types.InlineKeyboardButton(cell.value, callback_data=cell.value))
                else:
                    break

    if(my_report_list == []):
        my_report_list.append(telebot.types.InlineKeyboardButton("(Пусто)", callback_data="[none]"))                
    my_report_list.append(telebot.types.InlineKeyboardButton("-", callback_data="[none]"))
    my_report_list.append(telebot.types.InlineKeyboardButton("Вернуться", callback_data="[back_menu_my_reports]"))

    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*my_report_list)
    return inline_keyboard

#   Заход - Выход (В таблицу)
#   Функция чтения своей заявки
def Menu_ReportsMenu_MyReports_ReadReport():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy, user_link
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    call = copy_call_or_message
    wb = load_workbook(filename=file_path)
    worksheet = wb.active


    for row in worksheet.iter_rows(min_row=60, max_row=60):
        for cell in row:
            if cell.column > 2:
                if(str(cell.value) == call.data):

                    last_location = cell
                    cancel_report = telebot.types.InlineKeyboardButton("Удалить заявку", callback_data="[delete_my_report]")
                    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[back_my_report]")
                    back_to_my_report = [cancel_report, back]

                    Flag_my_problems = False    #   Отключаю поиск 
                    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
                    inline_keyboard.add(*back_to_my_report)
                    Del_All_Message()
                    message_del = bot.send_message(user_ID, f"{worksheet.cell(cell.row-3, cell.column).value}\n\n👤 Ответственный: {my_IT}\n\n💼 Proj-менеджер: {my_proj_men}\n\n🗣 Инициатор: {user_link}", reply_markup=inline_keyboard)
                    message_list.append(message_del.message_id)

                elif(cell.value is None):
                    break

    wb.close()

#   Заход - Выход (В таблицу)
#   Функция удаления своей заявки
def Menu_ReportsMenu_AllReports_ReadReport_DelReport():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    wb = load_workbook(filename=file_path)
    worksheet = wb.active
    ID_creator = ""
    cell_replace = ""
    id_report = ""
    for row in worksheet.iter_rows(min_row=60, max_row=60):
        for cell in row:
            if cell.column > 2:
                if(cell.value == last_location.value):
                    cell_replace = cell
                    id_report = worksheet.cell(60,cell.column).value
                    ID_creator = worksheet.cell(56, cell.column).value

                    ws.cell(row=56, column=cell.column, value="")   #   Удаляю ID User

                    ws.cell(row=57, column=cell.column, value="")   #   Удаляю описание проблемы

                    ws.cell(row=58, column=cell.column, value="")   #   Удаляю ID IT

                    ws.cell(row=59, column=cell.column, value="")   #   Удаляю ID Менеджера

                    ws.cell(row=60, column=cell.column, value="")   #   Удаляю ID Проблемы

                    break

    cell = cell_replace
    cell = worksheet.cell(cell.row, cell.column+1)

    while cell.value is not None:
        ws.cell(row=56, column=cell.column-1, value=worksheet.cell(56, cell.column).value)   #   Удаляю ID User
        ws.cell(row=57, column=cell.column-1, value=worksheet.cell(57, cell.column).value)   #   Удаляю описание проблемы
        ws.cell(row=58, column=cell.column-1, value=worksheet.cell(58, cell.column).value)   #   Удаляю ID IT
        ws.cell(row=59, column=cell.column-1, value=worksheet.cell(59, cell.column).value)   #   Удаляю ID Менеджера
        ws.cell(row=60, column=cell.column-1, value=worksheet.cell(60, cell.column).value)   #   Удаляю ID Проблемы

        
        ws.cell(row=56, column=cell.column, value="")   #   Удаляю ID User
        ws.cell(row=57, column=cell.column, value="")   #   Удаляю описание проблемы
        ws.cell(row=58, column=cell.column, value="")   #   Удаляю ID IT
        ws.cell(row=59, column=cell.column, value="")   #   Удаляю ID Менеджера
        ws.cell(row=60, column=cell.column, value="")   #   Удаляю ID Проблемы

        if cell.column == 100:
            break
        cell = worksheet.cell(cell.row, cell.column+1)
    
    wb.save(file_path)
    wb.close()
    Del_All_Message()   #   Удаляю все сообщения

    text = bot.send_message(ID_creator, f"🚫 Заявка \"{id_report}\" была закрыта!")
    bot.send_message(id_my_IT, f"🚫 Заявка \"{id_report}\" была удалена!")

    file = open('log.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text} - Кем: {user_ID}\n")
    else: file.write(f"{user_ID} - {now_time} - {text.text} - Кем: {user_ID}\n")
    file.close()

    message_del = bot.send_message(user_ID, f"📖 Выберите пункт 📖", reply_markup=Menu_ReportsMenu())
    message_list.append(message_del.message_id)

#   Функция середины создания заявки
def Menu_ReportsMenu_CreateNewReport_Start():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    Flag_problems = True

    crm = telebot.types.InlineKeyboardButton("CRM", callback_data="CRM")
    line = telebot.types.InlineKeyboardButton("Линия", callback_data="Линия")
    softphone = telebot.types.InlineKeyboardButton("Софтфон", callback_data="Софтфон")
    enother = telebot.types.InlineKeyboardButton("Другое", callback_data="Другое")
    
    list_but = [crm, line, softphone, enother]

    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*list_but)
    message_copy = bot.send_message(user_ID, "👇 Начало заведения заявки 👇")
    message_list.append(message_copy.message_id)
    message_del = bot.send_message(user_ID, "📍 Выберите тип проблемы:", reply_markup=inline_keyboard)

#   Функция начала создания заявки
def Menu_ReportsMenu_CreateNewReport_Mid():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    call = copy_call_or_message
    Flag_debug = True

    Kind = call.data   #   Передаю значение нажатой кнопки для определения типа проблемы
    message_copy = bot.send_message(user_ID, f"✅ Тип проблемы: {Kind}")
    message_list.append(message_copy.message_id)
    bot.delete_message(user_ID, message_del.message_id)   #   Удаляю прошло сообщение
    message = bot.send_message(user_ID, "📍 Введите ФИО:")   #   Запрашиваю ФИО
    last_message = message.text   #   Переменная для правильной отработки условий*
    echo_all(message)



###     Меню - Пролемы/Шаблоны/FAQ     ###



#   Заход - Выход (В таблицу)
#   Функция для работы с меню
def Menu_Movement(search_text):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV

    print("Open")
    wb = load_workbook(filename=file_path)
    worksheet = wb.active

    button_cells = []
    button_id = ""
    Flag = False

    
    for row in worksheet.iter_rows(min_row=4, max_row=4):  # Поиск по строкам
        if row[0].row == 4:
            for cell in row:
                if cell.value == search_text:
                    Flag = True
                    cell_copy = cell
                    break
                
    if(Flag):   # Вывод меню

        cell = cell_copy
        for i in range(12): # Создание списка кнопок
            cell_copy = worksheet.cell(cell.row+i+1, cell.column-1)
            button_id = str(worksheet.cell(cell.row+i+1, cell.column).value)

            if(button_id == "None"):    # Остановка при встрече пустых ячеек
                break
            button_cells.append(telebot.types.InlineKeyboardButton(cell_copy.value, callback_data=button_id))   # Добавление в список

        if(search_text == "[menu]" or search_text == "[back]"):

            Flag_my_problems = False    #   Флаг выключается, если ранее мы находились в просмотре своих заявок

            for y in acc_list:
                if(str(user_ID) == str(y)):
                    button_cells.append(telebot.types.InlineKeyboardButton("Заявки", callback_data="[reports]"))
            for y in admin_list:
                if(str(user_ID) == str(y)):
                    button_cells.append(telebot.types.InlineKeyboardButton("Администрирование", callback_data="[admin_menu]"))

                    
        print("Close")
        wb.close()
        last_button_cells = button_cells    # Переменная для вывода последнего меню
        inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
        inline_keyboard.add(*button_cells)
        return inline_keyboard
    
    
    else:   #   Вывод и поиск среди РЕШЕНИЙ

        for row in worksheet.iter_rows(min_row=21, max_row=21): # Поиск по строке
            for cell in row:
                if cell.value == str(search_text):
                    cell_copy = worksheet.cell(cell.row+1, cell.column).value
                    photo_copy = worksheet.cell(cell.row+2, cell.column).value
                    file_copy = worksheet.cell(cell.row+3, cell.column).value

                    if photo_copy is not None:
                        try:
                            image_paths = photo_copy.split('; ')
                            media = []  # добавлено
                            for path in image_paths:
                                with open(path, 'rb') as f:
                                    photo1 = f.read()
                                    if(image_paths[0] == path):    #    Проверка для установки caption=cell_copy, поскольку текст не будет выводиться
                                        media.append(telebot.types.InputMediaPhoto(photo1, caption=cell_copy))
                                    else:
                                        media.append(telebot.types.InputMediaPhoto(photo1))
                            
                            message_del = bot.send_media_group(user_ID, media)
                            try:
                                message_list.append(message_del[-1].message_id)
                                message_list.append(message_del[-2].message_id)
                            except:
                                message_list.append(message_del[-1].message_id)
                                
                        except:
                            bot.send_message(user_ID, "⚠️ Ошибка при выводе фото! Просьба сообщить IT. ⚠️\n📍 Код ошибки: 330")
                            message_del = bot.send_message(user_ID, cell_copy)
                            message_list.append(message_del.message_id)
                    else:
                        message_del = bot.send_message(user_ID, cell_copy)  # Вывод решения проблемы
                        message_list.append(message_del.message_id)

                    if file_copy is not None:
                        try:
                            with open(file_copy, 'rb') as file:
                                message_del = bot.send_document(user_ID, file)
                                message_list.append(message_del.message_id)
                        except:
                            bot.send_message(user_ID, "⚠️ Ошибка при выводе файла! Просьба сообщить IT. ⚠️\n📍 Код ошибки: 220")
                    
                    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
                    inline_keyboard.add(*last_button_cells)
                    return inline_keyboard
            if last_button_cells == []:
                bot.send_message(user_ID, "⚠️ Ошибка при выводе! Просьба сообщить IT. ⚠️\n📍 Код ошибки: 110")  # Вывод решения проблемы
                break
        print("Close")
        wb.close()    



###     Побочные функции (Подтверждения/Очистки/Переходы/Шаблоны)     ###



#   Заход - Выход (В таблицу)
#   Функция отправки новой заявки
def Send_New_Report():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    wb = load_workbook(filename=file_path)
    worksheet = wb.active

    report_id = random.randint(100001, 999999)   #   Переменная создаёт рандомный номер для ID заявки
    location_IT_ID = worksheet.cell(52,3).value    #   Переменная принимает значение ID IT из таблицы

    Del_All_Message()   #   Удаляю все сообщения
    
    bot.send_message(user_ID, f"✅ Заявка отправлена!\n💬 ID заявки - {report_id}\n\n📎 Заявка будет рассмотрена в ближайшее время, ожидайте!")    #   Вывод оповещения для пользователя
    text = f"✅ Заявка отправлена! 💬 ID заявки - {report_id}"

    file = open('log.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text}\n")
    else: file.write(f"{user_ID} - {now_time} - {text}\n")
    file.close()

    file = open('report_history.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    file.write(f"\nДата: {now_time}\nКем: {user_ID}\nID заявки: {report_id}\n{text_end_report}\n\n\n###  ###  ###\n\n")
    file.close()


    #   Выясняем proj-менеджера пользователя и его ссылку на telegram
    for row in worksheet.iter_rows(min_row=45, max_row=45):
        for cell in row:
            if cell.column > 2 and user_ID == cell.value:
                cell_1 = worksheet.cell(cell.row+2, cell.column).value
                my_proj_men = cell_1
                

    #   Добавление заявки в таблицу
    for row in worksheet.iter_rows(min_row=56, max_row=56):
        for cell in row:
            if cell.column > 2:
                if(cell.value is None):
                    ws.cell(row=56, column=cell.column, value=user_ID)   #   Вписываю ID User
                    ws.cell(row=57, column=cell.column, value=text_end_report)   #   Вписываю описание проблемы
                    ws.cell(row=58, column=cell.column, value=location_IT_ID)   #   Вписываю ID IT
                    ws.cell(row=59, column=cell.column, value=my_proj_men)   #   Вписываю ID Менеджера
                    ws.cell(row=60, column=cell.column, value=report_id)   #   Вписываю ID Проблемы
                    break
    wb.save(file_path)
    wb.close()

    bot.send_message(location_IT_ID, f"✅ Новая заявка!\n💬 ID заявки - {report_id}")   #   Вывод оповещения для IT о новой заявке

    #   Сброс переменных

    Flag_problems = False
    Kind = ""
    Team = "" 
    Name = ""
    Sip = ""
    Of_Ud = ""
    How_managed = ""
    decribtion = ""
    last_message = ""

    button_id = "[menu]"
    message_del = bot.send_message(user_ID, f"📖 Выберите пункт 📖", reply_markup=Menu_Movement(button_id))   #   Выходим в главное меню
    message_list.append(message_del.message_id)

#   Функция очистки всех переменных и выхода в меню
def Clear_Nums_And_Go_To_Menu():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV


    Flag_debug = False

    Flag_problems = False    #    Флаг, работает во время заведения заявки

    Flag_my_problems = False    #    Флаг, работает во время просмотра своих заявок

    Flag_add = False

    Flag_debug = False

    Kind = Team = Name = Sip = Of_Ud = How_managed = decribtion = ""    #    Переменные для составлени заявки

    last_message = ""    #    Переменная для правильной отработки условий*

    cell_copy = ""    #    Переменная для хранения cell в тех случаях, когда необходимо*

    copy_call_or_message = ""

    last_location = ""    #    Переменная для хранения местоположения при удалении заявки

    last_button_cells = ""

    text_end_report = ""   #   Переменная описания новой заявки

    location_last_empty = ""

    # Очистка переменных на создание заявки
    Team = ""
    Name = ""
    Sip = ""
    Of_Ud = ""
    How_managed = ""
    decribtion = ""
    my_proj_men = ""

    # Очистка переменных на добавление СВ / Администратора
    Name_SV = ""
    id_SV = ""
    link_SV = ""
    proj_SV = ""


    button_id = "[menu]"
    message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_Movement(button_id))
    Del_All_Message()   #   Удаляю все сообщения
    message_list.append(message_del.message_id)

#   Функция подтверждения о создании новой заявки
def Confirm_StartCreate_New_Report():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_report = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[create_report_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[create_report_cancel]")
    back_menu = [cancel_report, back]
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Подтверждение начала добавления нового СВ
def Confirm_Add_New_SV():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_report = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[add_SV_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[add_SV_cancel]")
    back_menu = [cancel_report, back]
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Подтверждение начала добавления нового Администратора 
def Confirm_Add_New_Admin():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_report = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[add_Admin_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[add_Admin_cancel]")
    back_menu = [cancel_report, back]
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Подтверждение начала удлаения СВ 
def Confirm_Del_New_SV():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_report = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[del_SV_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[del_SV_cancel]")
    back_menu = [cancel_report, back]
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Подтверждение начала удаления Администратора 
def Confirm_Del_New_Admin():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_report = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[del_Admin_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[del_Admin_cancel]")
    back_menu = [cancel_report, back]
    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Подтверждение окончания добавления нового СВ / Администратора 
def Confirm_Add_New_SV_or_Admin():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, copy_call_or_message, last_button_cells, location_last_empty
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV

    cancel_add = telebot.types.InlineKeyboardButton("Продолжить", callback_data="[add_continue]")
    back = telebot.types.InlineKeyboardButton("Вернуться", callback_data="[add_cancel]")

    back_menu = [cancel_add, back]

    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*back_menu)
    return inline_keyboard

#   Функция подтверждения отправки заявки
def Confirm_Send_New_Report(text_end_report):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    Flag_problems = True
    send_report = telebot.types.InlineKeyboardButton("✅ Отправить", callback_data="[send_report]")
    cancel_report = telebot.types.InlineKeyboardButton("⛔ Отменить", callback_data="[cancel_report]")
    
    list_but = [send_report, cancel_report]

    inline_keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    inline_keyboard.add(*list_but)

    message_del = bot.send_message(user_ID, f"💬 Итог:\n{text_end_report}", reply_markup=inline_keyboard)
    message_list.append(message_del.message_id)

#   Функция отмены отправки заявки
def Cancel_Send_Report():
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV

    #   Сброс переменных
    Flag_problems = False
    Kind = ""
    Team = "" 
    Name = ""
    Sip = ""
    Of_Ud = ""
    How_managed = ""
    decribtion = ""
    last_message = ""

    message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_ReportsMenu())   #   Выходим в главное меню
    Del_All_Message()   #   Удаляю все сообщения
    message_list.append(message_del.message_id)

#   Функция очистки всех сообщений
def Del_All_Message():
    global message_list, undel_mes
    if undel_mes != "":
        try:
            bot.delete_message(user_ID, undel_mes.message_id)   #   Удаляю прошло сообщение
            undel_mes = ""
        except:
            None
    for i in message_list:
        try:
            bot.delete_message(user_ID, i)   #   Удаляю прошло сообщение
        except:
            None
    message_list = []
                 
#   Функция разблокировки пользователя              
def Ban_ID_Del_From_BanList(target_id):
    lines = []
    found = False
    file_path = "ban.txt"

    with open(file_path, 'r') as file:
        for line in file:
            user_id = line.strip()
            if user_id != target_id:
                lines.append(line)  # Добавляем строку в список
            else:
                found = True

    if found:
        with open(file_path, 'w') as file:
            for line in lines:
                file.write(line)  # Записываем обновленные строки в файл
        ban_list_refresh()
        text = bot.send_message(user_ID, f'✅ Пользователь с ID {target_id} разблокирован!')
        bot.send_message(id_my_IT, f'✅ Пользователь с ID {target_id} разблокирован! Кем: {user_ID}')
        try:
            bot.send_message(target_id, f'✅ Вы были разблокированы!')
        except:
            None

        file = open('log.txt', 'a')
        now_time = datetime.datetime.now()
        now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
        if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
        else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
        file.close()
    else:
        bot.send_message(user_ID, f'⚠️ Пользователь с ID {target_id} не заблокирован!')

#   Функция блокировки пользователя  
def Ban_ID(target_id):
    file = open("ban.txt", 'a')
    file.write(f"{target_id}\n")
    file.close()
    ban_list_refresh()

    text = bot.send_message(user_ID, f'⛔️ Пользователь с ID {target_id} заблокирован.')
    bot.send_message(id_my_IT, f'⛔️ Пользователь с ID {target_id} был заблокирован. Кем: {user_ID}')
    try:
        bot.send_message(target_id, f'⛔️ Вы были заблокированы!')
    except:
        None


    file = open('log.txt', 'a')
    now_time = datetime.datetime.now()
    now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
    if(user_ID == ""): file.write(f"reload_sys - {now_time} - {text.text}\n")
    else: file.write(f"{user_ID} - {now_time} - {text.text}\n")
    file.close()

#   Функция-ШАБЛОН для ввода информации по новой заявке
def Input_Report_Sample(message, text_1, text_2):
    global message_list, undel_mes
    try:
        #   Удаляем прошлые сообщения
        bot.delete_message(user_ID, message.message_id-1)
        bot.delete_message(user_ID, message.message_id)
    except:
        None

    #   Выводим текст
    text = bot.send_message(user_ID, text_1)
    
    message_list.append(text.message_id)

    undel_mes = bot.send_message(user_ID, text_2)

#   Функция для ввода информации по новой заявке
def Input_Report(message):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug
    global admin_list, acc_list, message_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    if(message.text != last_message):
        
        if(Flag_add):
            if Name_SV == "":
                
                Name_SV = message.text   #   Присваиваем значение для Name
                text_1 = f"✅ ФИО: {Name_SV}"
                text_2 = "📍 Введите ID:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif id_SV == "":

                id_SV = message.text   #   Присваиваем значение для Team
                text_1 = f"✅ ID: {id_SV}"
                text_2 = "📍 Введите ссылку на TG:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif link_SV == "":

                link_SV = message.text   #   Присваиваем значение для Team
                text_1 = f"✅ Ссылка: {link_SV}"
                text_2 = "📍 Введите proj-менеджера этого СВ:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif proj_SV == "":

                message_list.append(message.message_id)
                Del_All_Message()   #   Удаляю все сообщения
                proj_SV = message.text   #   Присваиваем значение для Sip
                message_del = bot.send_message(user_ID, f"💬 Итог:\n\n✅ ФИО: {Name_SV}\n\n✅ ID: {id_SV}\n\n✅ Ссылка: {link_SV}\n\n✅ Proj-менеджер: {proj_SV}", reply_markup = Confirm_Add_New_SV_or_Admin())    #   Вывод оповещения для пользователя
                message_list.append(message_del.message_id)
                

                

        else:
                
            if Name == "":
                
                Name = message.text   #   Присваиваем значение для Name
                text_1 = f"✅ ФИО: {Name}"
                text_2 = "📍 Введите номер команды:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif Team == "":

                Team = message.text   #   Присваиваем значение для Team
                text_1 = f"✅ Номер команды: {Team}"
                text_2 = "📍 Введите актуальный SIP:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif Sip == "":

                Sip = message.text   #   Присваиваем значение для Sip
                text_1 = f"✅ SIP: {Sip}"
                text_2 = "📍 Офисный или удалённый сотрудник?"
                Input_Report_Sample(message, text_1, text_2)
                
            elif Of_Ud == "":

                Of_Ud = message.text   #   Присваиваем значение для Of_Ud
                text_1 = f"✅ Расположение: {Of_Ud}"
                text_2 = "📍 Опишите вашу проблему:"
                Input_Report_Sample(message, text_1, text_2)
                
            elif decribtion == "":
                decribtion = message.text   #   Присваиваем значение для decribtion
                text_1 = f"✅ Описание: \"{decribtion}\""
                text_2 = "📍 Что делали для решения проблемы?"
                Input_Report_Sample(message, text_1, text_2)
                
            elif How_managed == "":
                # Если имя еще не введено, сохраняем его в словаре
                How_managed = message.text
                
                del_list = [0,1,2,5,8,11,14,17,19]
                for i in del_list:
                    bot.delete_message(user_ID, message.message_id-i)
                
                Flag_debug = False
                text_end_report = f"\n✅ ФИО: {Name}\n\n✅ Номер команды: {Team}\n\n✅ SIP: {Sip}\n\n✅ Расположение: {Of_Ud}\n\n✅ Описание: \"{decribtion}\"\n\n✅ Что предприняли: \"{How_managed}\""
                Confirm_Send_New_Report(text_end_report)
                


###     Обработчики Нажатий/Ввода     ###



# Обработчик команды "/unban"
@bot.message_handler(commands=['unban'])
def unblock_user(message):
    global user_ID, message_del

    bot.delete_message(user_ID, message.message_id)
    # Разделяем сообщение на команду и аргументы
    command, user_id_del = message.text.split()
    Ban_ID_Del_From_BanList(user_id_del)

# Обработчик команды "/ban" 
@bot.message_handler(commands=['ban'])  
def block_user(message):
    global user_ID, message_del

    bot.delete_message(user_ID, message.message_id)
    # Разделяем сообщение на команду и аргументы
    command, user_id_del = message.text.split()
    Ban_ID(user_id_del)

# Обработчик нажатий
@bot.callback_query_handler(func=lambda call: True)
def button_callback(call): 
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy, for_user_link, user_link
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug, Flag_del_admin, Flag_del_SV
    global admin_list, acc_list, message_list, ban_list
    global Name_SV, id_SV, link_SV, proj_SV
    
    for_user_link = call.from_user.username
    user_link = f"https://t.me/{for_user_link}"
    copy_call_or_message = call
    user_ID = call.message.chat.id
    admin_list_refresh()
    acc_list_refresh()
    

    #  Проверка, находится ли пользователь в бане
    if str(user_ID) in str(ban_list):
        None
    else:
        file = open('log.txt', 'a')
        now_time = datetime.datetime.now()
        now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
        if(user_ID == ""): file.write(f"reload_sys - {now_time} - {call.data}\n")
        else: file.write(f"{user_ID} - {now_time} - {call.data}\n")
        file.close()

        #   Условие заведения заявки
        if(Flag_problems):

            #   Отмена заявки
            if(call.data == "[cancel_report]"):
                Cancel_Send_Report()
                
            #   Отправка заявки
            elif(call.data == "[send_report]"):
                Send_New_Report()

            #   Начало создания заявки
            else:
                Menu_ReportsMenu_CreateNewReport_Mid()

        #   Открытие меню Заявок
        elif(call.data == "[reports]"):  
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_ReportsMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[add_SV_continue]"):
            Del_All_Message()
            Flag_add_SV = True
            Flag_debug = True
            Flag_add = True
            bot.send_message(user_ID, "👇 Добавление нового супервайзра 👇\n\n📍 Введите ФИО супервайзера:")

        elif(call.data == "[add_Admin_continue]"):
            Del_All_Message()
            Flag_add_admin = True
            Flag_debug = True
            Flag_add = True
            bot.send_message(user_ID, "👇 Добавление нового администратора 👇\n\n📍 Введите ФИО администратора:")

        elif(call.data == "[add_continue]"):
            if(Flag_add_SV):
                Menu_AdminMenu_AddSV()
            else:
                Menu_AdminMenu_AddAdmin()

        elif(call.data == "[add_SV_cancel]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_AdminMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[add_Admin_cancel]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_AdminMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[del_SV_continue]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📍 Введите ID СВ:")
            message_list.append(message_del.message_id)
            Flag_del_SV = True
            

        elif(call.data == "[del_Admin_continue]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📍 Введите ID Администратора:")
            message_list.append(message_del.message_id)
            Flag_del_admin = True

        elif(call.data == "[del_Admin_cancel]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_AdminMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[del_SV_cancel]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_AdminMenu())
            message_list.append(message_del.message_id)
            
        elif(call.data == "[add_cancel]"):
            Del_All_Message()
            Clear_Nums_And_Go_To_Menu()

        elif(call.data == "[admin_menu]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_AdminMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[add_SV]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "⚠️ Вы уверены, что хотите добавить нового СВ?", reply_markup=Confirm_Add_New_SV())
            message_list.append(message_del.message_id)

        elif(call.data == "[add_admin]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "⚠️ Вы уверены, что хотите добавить нового Администратора?", reply_markup=Confirm_Add_New_Admin())
            message_list.append(message_del.message_id)


        elif(call.data == "[del_SV]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "⚠️ Вы уверены, что хотите удалить СВ?", reply_markup=Confirm_Del_New_SV())
            message_list.append(message_del.message_id)

        elif(call.data == "[del_admin]"):
            Del_All_Message()
            message_del = bot.send_message(user_ID, "⚠️ Вы уверены, что хотите удалить Администратора?", reply_markup=Confirm_Del_New_Admin())
            message_list.append(message_del.message_id)

        elif(call.data == "[take_link]"):
            Menu_Enother_TakeLink()

        elif(call.data == "[take_id]"):
            Menu_Enother_TakeID()
            
        elif(call.data == "[delete_my_report]"):
            Menu_ReportsMenu_AllReports_ReadReport_DelReport()

        #   Возврат к основному меню из меню "Мои заявки"
        elif(call.data == "[back_menu_my_reports]"):
            Del_All_Message()   #   Удаляю все сообщения
            message_del = bot.send_message(user_ID, f"📖 Выберите пункт 📖", reply_markup=Menu_ReportsMenu())
            message_list.append(message_del.message_id)

        elif(call.data == "[my_report]" or call.data == "[back_my_report]"):
            Del_All_Message()   #   Удаляю все сообщения
            message_del = bot.send_message(user_ID, f"📖 Мои заявки 📖", reply_markup=Menu_ReportsMenu_MyReports())
            message_list.append(message_del.message_id)

        elif (call.data == "[all_reports]"):
            Del_All_Message()   #   Удаляю все сообщения
            message_del = bot.send_message(user_ID, f"📖 Все заявки 📖", reply_markup=Menu_ReportsMenu_AllReports())
            message_list.append(message_del.message_id)

        elif (call.data == "[create_report]" or call.data == "[create_report_continue]" or call.data == "[create_report_cancel]"):
            
            if(call.data == "[create_report]"):
                Del_All_Message()   #   Удаляю все сообщения
                message_del = bot.send_message(user_ID, "⚠️ Уверены, что хотите составить заявку по тех. проблеме?\n\n📍 Заявки в выходные дни не обрабатываются.\n\n📍 Вы соглашаетесь с тем, что выполнили все указанные действия по решению проблемы, представленные во вкладке \"Меню\".\n\n💬 Необходимо будет ввести данные:\n- ФИО\n- № команды\n- SIP\n- Офис/Удалённый\n- Описание проблемы\n\n📎 Заявка будет рассмотрена в ближайшие 5-10 минут.", reply_markup=Confirm_StartCreate_New_Report())   # После нажатия ищу меню
                message_list.append(message_del.message_id)

            elif(call.data == "[create_report_continue]"):
                Del_All_Message()   #   Удаляю все сообщения
                Menu_ReportsMenu_CreateNewReport_Start()

            elif(call.data == "[create_report_cancel]"):
                Del_All_Message()   #   Удаляю все сообщения
                message_del = bot.send_message(user_ID, f"📖 Выберите пункт 📖", reply_markup=Menu_ReportsMenu())
                message_list.append(message_del.message_id)

        #  Возврат к основному меню из меню "Мои заявки"
        elif(call.data == "[back_menu]"):
            Del_All_Message()   #   Удаляю все сообщения
            button_id = "[menu]"
            message_del = bot.send_message(user_ID, f"📖 Выберите пункт 📖", reply_markup=Menu_Movement(button_id))
            message_list.append(message_del.message_id)

        #  Чтение своей заявки
        elif(Flag_my_problems):
            Menu_ReportsMenu_MyReports_ReadReport()

        #  Перемещение по меню
        else:
            #   Проверка на нажатие на черточку
            if(call.data != '[none]'):
                Del_All_Message()   #   Удаляю все сообщения
                message_del = bot.send_message(user_ID, "📖 Выберите пункт 📖", reply_markup=Menu_Movement(call.data))   # После нажатия ищу меню
                message_list.append(message_del.message_id)
                
# Обработчик ввода текстовых сообщений
@bot.message_handler(func=lambda message: True)
def echo_all(message):
    global message_del, Kind, user_ID, last_message, last_location, file_path, ws, wb, copy_call_or_message, last_button_cells, location_last_empty, id_my_IT, cell_copy, for_user_link
    global Team, Name, Sip, Of_Ud, How_managed, decribtion, text_end_report, my_proj_men
    global Flag_problems, Flag_my_problems, Flag_add_SV, Flag_add_admin, Flag_add, Flag_debug, Flag_del_SV, Flag_del_admin
    global admin_list, acc_list, message_list, ban_list
    global Name_SV, id_SV, link_SV, proj_SV
    

    current_time = time.time()

    message_global = message
    user_ID = message.chat.id
    for_user_link = message.from_user.username

    #  Проверка, находится ли пользователь в бане
    if str(user_ID) in str(ban_list):
        None
    else:

        file = open('log.txt', 'a')
        now_time = datetime.datetime.now()
        now_time = now_time.strftime("%d/%m/%y - %H:%M:%S")
        if(user_ID == ""): file.write(f"reload_sys - {now_time} - {message.text}\n")
        else: file.write(f"{user_ID} - {now_time} - {message.text}\n")
        file.close()

        
        if(message.text == "/start"):
            message_list.append(message_global.message_id)   #   Записываю ID сообщения
            Del_All_Message()   #   Удаляю все сообщения
            Clear_Nums_And_Go_To_Menu()

        elif Flag_del_SV:
            Menu_AdminMenu_DelSV(message.text)

        elif Flag_del_admin:
            Menu_AdminMenu_DelAdmin(message.text)

        elif Flag_debug:
            if user_ID in last_message_times and current_time - last_message_times[user_ID] < 1:
                message_list.append(message.message_id)   #   Записываю ID сообщения
                bot.send_message(user_ID, "⚠️ Вы печатаете слишком быстро!\n⚠️ Возврат в начальное меню!")

                Del_All_Message()   #   Удаляю все сообщения
                Clear_Nums_And_Go_To_Menu()
                
            else:
                if(Flag_problems):
                    Input_Report(message)
                    last_message_times[user_ID] = current_time
                elif Flag_add:
                    Input_Report(message)
                    last_message_times[user_ID] = current_time
                else:
                    send_inline(message)
                    last_message_times[user_ID] = current_time
        else:
            try:
                bot.delete_message(user_ID, message.message_id)
                message_warning = bot.send_message(message.chat.id, "⚠️ Для старта введите \"/start\" ⚠️")
                message_list.append(message_warning.message_id)   #   Записываю ID сообщения
                    

            except:
                message_list.append(message_global.message_id)   #   Записываю ID сообщения
                message_warning = bot.send_message(message.chat.id, "⚠️ Для старта введите \"/start\" ⚠️")

# Обработчик команды "/start"
@bot.message_handler(commands=['start'])
def send_inline(message):
    global Flag_problems, message_del, user_ID, Flag_problems, ban_list   #   Инициализирую глоб. переменные

    #  Проверка, находится ли пользователь в бане
    if str(user_ID) in str(ban_list):
        None
    else:
        user_ID = message.chat.id
        Clear_Nums_And_Go_To_Menu()

bot.polling()